"""
Script: evaluation_sheet_generator_v2.py
Description: Generates an Excel spreadsheet for human evaluation of image captions.
             Randomly selects between greedy and beam search caption generation.
             The 'Generation Method' column is now hidden by default.
"""

import os
import random
import logging
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font
from tqdm import tqdm

from vtt.data.caption_preprocessing import load_and_clean_captions, load_tokenizer
from vtt.config import START_TOKEN, END_TOKEN
from vtt.utils import set_seed
from vtt.data.data_loader import load_split_datasets
from vtt.models.predict import generate_caption_greedy, generate_caption_beam
from vtt.models.decoder import build_decoder_model
from PIL import Image as PILImage
import tensorflow as tf


# --- Logging Setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M",
)
logger = logging.getLogger(__name__)


# --- Main ---
def main():
    # Set random seed for repeatability
    set_seed(42)
    # Set number of evaluation samples to include
    NUM_EVAL_SAMPLES = 100
    # Define beam width for beam search (can be adjusted)
    BEAM_WIDTH = 5

    current_script_dir, project_root, data_dir, processed_dir, raw_dir, model_dir = (
        get_project_paths()
    )

    # Note: The images directory path needs to be the user's local path to the images
    # since the images are not part of the repository.
    image_dir = "/mnt/c/grad_school/northeastern/ie7374/project/data/flickr8k_images"
    captions_csv = os.path.join(raw_dir, "flickr8k_captions.csv")
    features_npz = os.path.join(processed_dir, "flickr8k_features.npz")
    captions_npz = os.path.join(processed_dir, "flickr8k_padded_caption_sequences.npz")
    tokenizer_path = os.path.join(processed_dir, "flickr8k_tokenizer.json")
    model_weights = os.path.join(model_dir, "flickr8k_decoder_weights.weights.h5")
    output_file = os.path.join(current_script_dir, "human_evaluation_sheet_v2.xlsx")

    model, tokenizer, max_len = load_model_and_tokenizer(
        tokenizer_path, model_weights, features_npz, captions_npz
    )
    captions_map = load_and_clean_captions(captions_csv)

    _, _, (test_features, _, test_ids) = load_split_datasets(
        features_path=features_npz, captions_path=captions_npz, return_numpy=True
    )

    selected_ids, selected_features = select_evaluation_samples(
        test_ids, test_features, NUM_EVAL_SAMPLES
    )

    evaluation_data = generate_evaluation_data(
        selected_ids,
        selected_features,
        model,
        tokenizer,
        max_len,
        captions_map,
        BEAM_WIDTH,
    )

    create_evaluation_spreadsheet(image_dir, evaluation_data, output_file)
    logger.info("--- Evaluation Sheet Generation Complete ---")


# --- Helper Functions ---
def get_project_paths() -> Tuple[str, str, str, str, str, str]:
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(current_script_dir, "..", ".."))
    data_dir = os.path.join(project_root, "data")
    processed_dir = os.path.join(data_dir, "processed")
    raw_dir = os.path.join(data_dir, "raw")
    model_dir = os.path.join(project_root, "models")
    return current_script_dir, project_root, data_dir, processed_dir, raw_dir, model_dir


def load_model_and_tokenizer(
    tokenizer_path: str, model_weights_path: str, features_path: str, captions_path: str
):
    logger.info("Loading tokenizer...")
    tokenizer = load_tokenizer(tokenizer_path)
    vocab_size = tokenizer.num_words

    logger.info("Determining maximum caption length...")
    train_ds, _, _ = load_split_datasets(
        features_path=features_path,
        captions_path=captions_path,
        batch_size=64,
        val_split=0.15,
        test_split=0.10,
        shuffle=True,
        buffer_size=1000,
        seed=42,
        cache=True,
        return_numpy=False,
    )
    for _, input_caption, _ in train_ds.take(1):
        max_len = input_caption.shape[1]
        break

    logger.info("Building model...")
    model = build_decoder_model(vocab_size=vocab_size, max_caption_len=max_len)
    model.load_weights(model_weights_path)
    logger.info("Model loaded successfully.")

    return model, tokenizer, max_len


def select_evaluation_samples(image_ids: List[str], image_features, num_samples: int):
    if len(image_ids) > num_samples:
        logger.info(f"Selecting {num_samples} random samples...")
        indices = random.sample(range(len(image_ids)), num_samples)
        return [image_ids[i] for i in indices], image_features[indices]
    else:
        logger.info(f"Using all {len(image_ids)} test images.")
        return image_ids, image_features


def generate_evaluation_data(
    image_ids, image_features, model, tokenizer, max_len, captions_map, beam_width: int
):
    data = []
    # Define generation methods and their names
    generation_methods = [
        (generate_caption_greedy, "Greedy"),
        (generate_caption_beam, "Beam Search"),
    ]

    for i, img_id in tqdm(
        enumerate(image_ids), total=len(image_ids), desc="Generating Captions"
    ):
        gt_list = captions_map.get(img_id, [])
        gt_caption = random.choice(gt_list) if gt_list else "N/A"
        gt_caption = gt_caption.replace(START_TOKEN, "").replace(END_TOKEN, "").strip()

        # Randomly choose a generation method
        chosen_method_func, method_name = random.choice(generation_methods)

        if method_name == "Greedy":
            caption = chosen_method_func(model, tokenizer, image_features[i], max_len)
        else:  # Beam Search
            caption = chosen_method_func(
                model, tokenizer, image_features[i], max_len, beam_width
            )

        caption = caption.replace(START_TOKEN, "").replace(END_TOKEN, "").strip()

        data.append(
            {
                "image_file": img_id,
                "ground_truth_caption": gt_caption,
                "generated_caption": caption,
                "generation_method": method_name,
            }
        )
    return data


def resize_image_for_excel(
    image_path: str, max_width: int = 250, max_height: int = 180
) -> ExcelImage:
    img = ExcelImage(image_path)
    pil_img = PILImage.open(image_path)
    orig_w, orig_h = pil_img.size

    if orig_w > max_width:
        new_h = int(orig_h * (max_width / orig_w))
        img.width = max_width
        img.height = new_h
    else:
        img.width = orig_w
        img.height = orig_h

    if img.height > max_height:
        img.width = int(img.width * (max_height / img.height))
        img.height = max_height

    return img


def create_evaluation_spreadsheet(
    image_dir, captions_data, output_filename="human_evaluation.xlsx"
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Captions Evaluation"

    # Important: The order of headers determines column letters
    headers = [
        "Image",  # A
        "Image Filename",  # B
        "Generated Caption (Model)",  # C
        "Adequacy (1-5)",  # D
        "Fluency (1-5)",  # E
        "Overall Quality (1-5)",  # F
        "Comments",  # G
        "Generation Method",  # H
        "Ground Truth Caption",  # I
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = bold_font

    # Define column widths
    column_widths = [
        40,  # Image(A)
        30,  # Filename(B)
        40,  # Generated(C)
        20,  # Adequacy(D)
        20,  # Fluency(E)
        20,  # Overall Quality(F)
        40,  # Comments(G)
        20,  # Generation Method(H)
        40,  # Ground Truth(I)
    ]
    for col_idx, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + col_idx)].width = width

    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Entry",
        error="Enter a whole number (1–5)",
        showInputMessage=True,
        promptTitle="Rating",
        prompt="Enter a number from 1 to 5",
    )

    row = 2
    for item in captions_data:
        image_path = os.path.join(image_dir, item["image_file"])
        if not os.path.exists(image_path):
            logger.warning(f"Missing image: {image_path}")
            continue

        try:
            img = resize_image_for_excel(image_path)
            ws.row_dimensions[row].height = img.height / 0.75
            ws.add_image(img, f"A{row}")

            ws[f"B{row}"].value = item["image_file"]
            ws[f"C{row}"].value = item["generated_caption"]
            ws[f"D{row}"].value = ""  # Adequacy
            ws[f"E{row}"].value = ""  # Fluency
            ws[f"F{row}"].value = ""  # Overall Quality
            ws[f"G{row}"].value = ""  # Comments
            ws[f"H{row}"].value = item["generation_method"]
            ws.column_dimensions["H"].hidden = True
            ws[f"I{row}"].value = item["ground_truth_caption"]
            ws.column_dimensions["I"].hidden = True

            # Apply text wrapping for appropriate columns
            for col in "BCGHI":
                ws[f"{col}{row}"].alignment = Alignment(wrapText=True, vertical="top")

            # Apply data validation for rating columns (D, E, F)
            for col in "DEF":
                dv.add(f"{col}{row}")

            row += 1

        except Exception as e:
            logger.error(
                f"Failed to process image {item['image_file']}: {e}", exc_info=True
            )
            continue

    wb.save(output_filename)
    logger.info(f"Spreadsheet saved to {output_filename}")


if __name__ == "__main__":
    main()
